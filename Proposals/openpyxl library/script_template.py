
from time import time
import datetime

import os
import csv

import openpyxl
from openpyxl.styles import colors
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.cell import get_column_letter

import requests
import json

inputFileName = 'export-228-Hire-8600-2015-08-26-2.csv'
outputFileName = 'Output File (var. 9).xlsx'

login = 'speedio'
password = 'spd00ii99'
url = 'http://45.55.39.32:3000/cpf'

def readerInputWB(name):
    data = []
    with open(name, 'rb') as csvfile:
        spamreader = csv.reader(csvfile, delimiter=',')
        for row in spamreader:
            data += [row]

    return data


def formatCell():
    pass


def formatFone(fone):
    fone = '('+fone[:2]+')'+fone[2:]  # to add '()'
    fone = fone[:-4]+'-'+fone[-4:]  # to add '-'

    return fone


def formatterWB(data):
    # create new workbook
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "sheet1"

    ws2 = wb.create_sheet()
    ws2.title = "sheet2"
    
    nrow = len(data)
    ncol = 20  # len(data[0])

    quantityPhoneNumber = 0

    # set colors of rows
    verySoftCyanFill = PatternFill(start_color='95E6E9', end_color='95E6E9', fill_type='solid')
    veryLightCyanFill = PatternFill(start_color='E4F9FA', end_color='E4F9FA', fill_type='solid')
    
    ii = 0
    for i in range(1, nrow):
        ii += 1

        # pdfdata__date_of_activity
        if data[i][19]:
            try:
                date = datetime.datetime.strptime(data[i][19], '%Y-%m-%d %H:%M:%S').date()
                ws1.cell(row=ii+5, column=1).value = date.strftime('%d-%b-%y')
            except ValueError:
                print i
                pass

            try:
                date = datetime.datetime.strptime(data[i][19], '%m/%d/%y %H:%M %p').date()
                ws1.cell(row=ii+5, column=1).value = date.strftime('%d-%b-%y')
            except ValueError:
                pass

        else:
            ws1.cell(row=ii+5, column=1).value = ''

        # Razao Social
        ws1.cell(row=ii+5, column=2).value = data[i][7]

        # CNPJ
        ws1.cell(row=ii+5, column=3).value = data[i][8]

        # Classificacao
        ws1.cell(row=ii+5, column=4).value = data[i][9]

        # Capital Social
        ws1.cell(row=ii+5, column=5).value = data[i][17]

        # Nome
        ws1.cell(row=ii+5, column=6).value = str(data[i][3]).replace(',', '\n').replace("'", '')

        # CPF

        #
        ws1.cell(row=ii+5, column=8).value = str(data[i][6]).replace(' / ', '\n').replace('(', '0').replace(') ', '').replace('-', '')

        #
        ws1.cell(row=ii+5, column=9).value = data[i][5]

        # Rua
        ws1.cell(row=ii+5, column=12).value = data[i][10]

        # Numero
        ws1.cell(row=ii+5, column=13).value = data[i][13]

        # Bairro
        ws1.cell(row=ii+5, column=14).value = data[i][11]

        # Complemento
        ws1.cell(row=ii+5, column=15).value = data[i][14]

        # CEP
        ws1.cell(row=ii+5, column=16).value = data[i][15]

        # Cidade
        ws1.cell(row=ii+5, column=17).value = data[i][12]

        # Estado
        ws1.cell(row=ii+5, column=18).value = data[i][16]

        # Tipo
        ws1.cell(row=ii+5, column=19).value = data[i][20]

        # Fone
        cpf = str(data[i][2]).replace("'", '').replace(".", '').replace("-", '').split(',')
        payload = {'login': login, 'password': password, 'cpf': cpf}
        r = requests.post(url, data=payload)

        print('cpf numbers: ' + str(cpf))
        print('phone numbers: ' + str([r.json()[index]['fone'] for index in range(len(r.json()))]))

        print len(r.json())

        if len(r.json()) == 0:
            pass

        else:
            for numberFone in range(len(r.json())):
                ws1.cell(row=ii+5, column=20).value = formatFone(r.json()[numberFone]['fone'])

                quantityPhoneNumber += 1
                ws2.cell(row=quantityPhoneNumber, column=1).value = formatFone(r.json()[numberFone]['fone'])
                ws2.cell(row=quantityPhoneNumber, column=2).value = r.json()[numberFone]['fone']

                if numberFone == len(r.json())-1:
                    pass

                else:
                    # set alignment, fill and font
                    for j in range(ncol):
                        if i % 2 == 0:
                            ws1.cell(row=ii+5, column=j+1).fill = verySoftCyanFill
                        else:
                            ws1.cell(row=ii+5, column=j+1).fill = veryLightCyanFill
                        ws1.cell(row=ii+5, column=j+1).font = Font(name='Helvetica', size=10, bold=False, color=colors.BLACK)
                        ws1.cell(row=ii+5, column=j+1).alignment = Alignment(horizontal='left', vertical='top')

                    ii += 1

        # set alignment, fill and font
        for j in range(ncol):
            if i % 2 == 0:
                ws1.cell(row=ii+5, column=j+1).fill = verySoftCyanFill
            else:
                ws1.cell(row=ii+5, column=j+1).fill = veryLightCyanFill
            ws1.cell(row=ii+5, column=j+1).font = Font(name='Helvetica', size=10, bold=False, color=colors.BLACK)
            ws1.cell(row=ii+5, column=j+1).alignment = Alignment(horizontal='left', vertical='top')

    # set column width
    columnWidths = []
    for i in range(ii):
        for j, cell in enumerate([ws1.cell(row=i+5, column=j+1).value for j in range(ncol)]):
            if type(cell) != unicode:
                cell = str(cell)

            if cell.count('\n') != 0:
                cell = cell[:cell.index('\n')]

            if len(columnWidths) <= j:
                columnWidths += [len(cell)]
            else:
                if len(cell) > columnWidths[j]:
                    columnWidths[j] = len(cell)

    for j in range(ncol):
        ws1.column_dimensions[get_column_letter(j+1)].width = columnWidths[j]

    # return workbook
    return wb

def writerOutputWB(wb, name):
    wb.save(name)


def main(inputFileName, outputFileName):
    data = readerInputWB(inputFileName)
    wb = formatterWB(data)
    writerOutputWB(wb, outputFileName)


if __name__ == '__main__':
    start = time()
    main(inputFileName, outputFileName)
    stop = time()

    print('time: ' + str(stop - start))